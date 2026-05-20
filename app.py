"""
ISKCON Book Store - Complete Flask Application
=============================================
A production-ready e-commerce platform for ISKCON books.
"""

import os
import uuid
import hmac
import hashlib
import json
import csv
import io
import base64
import zipfile
from datetime import datetime, timedelta
from functools import wraps

# Load .env file for local development (no-op on Railway/Render where vars are injected)
from dotenv import load_dotenv
load_dotenv()

from flask import (
    Flask, render_template, request, session, redirect,
    url_for, flash, jsonify, abort, send_from_directory
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import or_
from email_utils import send_order_confirmation, send_order_shipped, send_order_delivered

# ─────────────────────────────────────────────
# App & Configuration
# ─────────────────────────────────────────────

app = Flask(__name__)

_IST = timedelta(hours=5, minutes=30)

@app.template_filter('ist')
def to_ist(dt, fmt='%d %b %Y, %I:%M %p'):
    if dt is None:
        return ''
    return (dt + _IST).strftime(fmt)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


class Config:
    SECRET_KEY = os.environ.get("SECRET_KEY", "iskcon-books-super-secret-key-2024")
    SQLALCHEMY_DATABASE_URI = os.environ.get("DATABASE_URL", f"sqlite:///{os.path.join(BASE_DIR, 'iskcon_books.db')}")
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "images", "books")
    MAX_CONTENT_LENGTH = 16 * 1024 * 1024          # 16 MB
    ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
    EBOOK_FOLDER = os.path.join(BASE_DIR, "ebooks")
    PREVIEW_FOLDER = os.path.join(BASE_DIR, "static", "previews")
    ALLOWED_EBOOK_EXTENSIONS = {"pdf", "epub"}
    RAZORPAY_KEY_ID     = os.environ.get("RAZORPAY_KEY_ID", "rzp_test_your_key_id")
    RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "your_razorpay_secret")
    PAYU_MERCHANT_KEY   = os.environ.get("PAYU_MERCHANT_KEY", "")
    PAYU_MERCHANT_SALT  = os.environ.get("PAYU_MERCHANT_SALT", "")
    PAYU_ENV            = os.environ.get("PAYU_ENV", "test")   # "test" or "prod"
    ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
    ADMIN_PASSWORD_HASH = generate_password_hash(os.environ.get("ADMIN_PASSWORD", "Hare@Krishna108"))
    WHATSAPP_NUMBER = os.environ.get("WHATSAPP_NUMBER", "+919999999999")
    UPI_ID   = os.environ.get("UPI_ID", "")
    UPI_NAME = os.environ.get("UPI_NAME", "ISKCON Book Store")
    STORE_NAME = "ISKCON Book Store"
    SHIPPING_CHARGE = float(os.environ.get("SHIPPING_CHARGE", "50"))
    FREE_SHIPPING_ABOVE = float(os.environ.get("FREE_SHIPPING_ABOVE", "500"))
    DELHIVERY_API_TOKEN       = os.environ.get("DELHIVERY_API_TOKEN", "")
    DELHIVERY_DEFAULT_WEIGHT  = float(os.environ.get("DELHIVERY_DEFAULT_WEIGHT", "0.1"))
    MAIL_USERNAME = os.environ.get("MAIL_USERNAME", "noreply@iskconbooks.in")
    BREVO_API_KEY = os.environ.get("BREVO_API_KEY", "")


app.config.from_object(Config)

# Fix Heroku/Render postgres:// → postgresql://
db_url = app.config["SQLALCHEMY_DATABASE_URI"]
print(f"[DB] Using: {db_url[:40]}...")  # debug — shows first 40 chars only
if db_url.startswith("postgres://"):
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url.replace("postgres://", "postgresql://", 1)

db = SQLAlchemy(app)

try:
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    os.makedirs(app.config["EBOOK_FOLDER"], exist_ok=True)
    os.makedirs(app.config["PREVIEW_FOLDER"], exist_ok=True)
except Exception as e:
    print(f"[WARNING] makedirs failed: {e}")


# ─────────────────────────────────────────────
# Database Models
# ─────────────────────────────────────────────

class Category(db.Model):
    __tablename__ = "categories"
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(100), nullable=False, unique=True)
    slug        = db.Column(db.String(100), nullable=False, unique=True)
    description = db.Column(db.Text)
    icon        = db.Column(db.String(10), default="📚")
    sort_order  = db.Column(db.Integer, default=0)
    books       = db.relationship("Book", backref="category", lazy=True)

    def __repr__(self):
        return f"<Category {self.name}>"


class Book(db.Model):
    __tablename__ = "books"
    id             = db.Column(db.Integer, primary_key=True)
    title          = db.Column(db.String(250), nullable=False)
    author         = db.Column(db.String(200), nullable=False)
    description    = db.Column(db.Text)
    short_desc     = db.Column(db.String(300))
    price          = db.Column(db.Float, nullable=False)
    original_price = db.Column(db.Float)
    image          = db.Column(db.String(200), default="default_book.jpg")
    category_id    = db.Column(db.Integer, db.ForeignKey("categories.id"))
    isbn           = db.Column(db.String(30))
    language       = db.Column(db.String(50), default="English")
    pages          = db.Column(db.Integer)
    weight_kg      = db.Column(db.Float, default=0.1)   # shipping weight
    publisher      = db.Column(db.String(200), default="The Bhaktivedanta Book Trust")
    stock          = db.Column(db.Integer, default=100)
    featured       = db.Column(db.Boolean, default=False)
    active         = db.Column(db.Boolean, default=True)
    deleted        = db.Column(db.Boolean, default=False)   # True = moved to Trash
    is_ebook       = db.Column(db.Boolean, default=False)
    ebook_file     = db.Column(db.String(200), nullable=True)
    preview_file   = db.Column(db.String(200), nullable=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    order_items    = db.relationship("OrderItem", backref="book", lazy=True)

    @property
    def discount_percent(self):
        if self.original_price and self.original_price > self.price:
            return int((1 - self.price / self.original_price) * 100)
        return 0

    @property
    def in_stock(self):
        return self.stock > 0

    def __repr__(self):
        return f"<Book {self.title}>"


class Order(db.Model):
    __tablename__ = "orders"
    id                 = db.Column(db.Integer, primary_key=True)
    order_number       = db.Column(db.String(20), unique=True, nullable=False)
    customer_name      = db.Column(db.String(200), nullable=False)
    customer_email     = db.Column(db.String(200))
    customer_phone     = db.Column(db.String(20), nullable=False)
    address            = db.Column(db.Text, nullable=False)
    city               = db.Column(db.String(100))
    state              = db.Column(db.String(100))
    pincode            = db.Column(db.String(10))
    subtotal           = db.Column(db.Float, nullable=False)
    shipping_charge    = db.Column(db.Float, default=0)
    discount_amount    = db.Column(db.Float, default=0)
    total_amount       = db.Column(db.Float, nullable=False)
    payment_method     = db.Column(db.String(50), default="cod")
    payment_status     = db.Column(db.String(50), default="pending")   # pending/paid/failed
    order_status       = db.Column(db.String(50), default="placed")    # placed/confirmed/shipped/delivered/cancelled
    razorpay_order_id  = db.Column(db.String(100))
    razorpay_payment_id = db.Column(db.String(100))
    coupon_code        = db.Column(db.String(50))
    notes              = db.Column(db.Text)
    created_at         = db.Column(db.DateTime, default=datetime.utcnow)
    courier_name       = db.Column(db.String(100))
    tracking_number    = db.Column(db.String(100))
    expected_delivery  = db.Column(db.Date)
    upi_transaction_id = db.Column(db.String(100))
    is_deleted         = db.Column(db.Boolean, default=False)
    customer_id        = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=True)
    items              = db.relationship("OrderItem", backref="order", lazy=True)

    def __repr__(self):
        return f"<Order {self.order_number}>"


class OrderItem(db.Model):
    __tablename__ = "order_items"
    id         = db.Column(db.Integer, primary_key=True)
    order_id   = db.Column(db.Integer, db.ForeignKey("orders.id"), nullable=False)
    book_id    = db.Column(db.Integer, db.ForeignKey("books.id"))
    book_title = db.Column(db.String(250))
    quantity   = db.Column(db.Integer, nullable=False)
    price      = db.Column(db.Float, nullable=False)

    @property
    def subtotal(self):
        return self.price * self.quantity


class Coupon(db.Model):
    __tablename__ = "coupons"
    id             = db.Column(db.Integer, primary_key=True)
    code           = db.Column(db.String(50), unique=True, nullable=False)
    description    = db.Column(db.String(200))
    discount_type  = db.Column(db.String(20), default="percent")   # percent / fixed
    discount_value = db.Column(db.Float, nullable=False)
    min_order      = db.Column(db.Float, default=0)
    max_discount   = db.Column(db.Float)                            # cap for percent coupons
    max_uses       = db.Column(db.Integer, default=100)
    used_count     = db.Column(db.Integer, default=0)
    active         = db.Column(db.Boolean, default=True)
    expires_at     = db.Column(db.DateTime)

    def is_valid(self, cart_total):
        if not self.active:
            return False, "Coupon is inactive."
        if self.used_count >= self.max_uses:
            return False, "Coupon usage limit reached."
        if self.expires_at and datetime.utcnow() > self.expires_at:
            return False, "Coupon has expired."
        if cart_total < self.min_order:
            return False, f"Minimum order ₹{self.min_order:.0f} required."
        return True, "Valid"

    def calculate_discount(self, cart_total):
        if self.discount_type == "percent":
            disc = cart_total * self.discount_value / 100
            if self.max_discount:
                disc = min(disc, self.max_discount)

            return round(disc, 2)
        return min(self.discount_value, cart_total)


class StockReceipt(db.Model):
    """Records each batch of books received from the ISKCON temple main store."""
    __tablename__ = "stock_receipts"
    id             = db.Column(db.Integer, primary_key=True)
    book_id        = db.Column(db.Integer, db.ForeignKey("books.id"), nullable=True)
    book_name      = db.Column(db.String(250), nullable=False)   # stored in case book is deleted
    quantity       = db.Column(db.Integer, nullable=False)
    cost_per_unit  = db.Column(db.Float, nullable=False)         # price paid to temple per copy
    total_payment  = db.Column(db.Float, nullable=False)         # quantity × cost_per_unit
    payment_status = db.Column(db.String(20), default="paid")    # paid / pending
    received_date  = db.Column(db.DateTime, default=datetime.utcnow)
    notes          = db.Column(db.Text)
    batch_ref      = db.Column(db.String(20), nullable=True)     # shared ID for bulk-upload rows
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    book           = db.relationship("Book", backref="stock_receipts", lazy=True)


class Setting(db.Model):
    __tablename__ = "settings"
    key   = db.Column(db.String(100), primary_key=True)
    value = db.Column(db.String(500), nullable=True)

    @staticmethod
    def get(key, default=None):
        s = Setting.query.get(key)
        return s.value if s else default

    @staticmethod
    def set(key, value):
        s = Setting.query.get(key)
        if s:
            s.value = value
        else:
            db.session.add(Setting(key=key, value=value))
        db.session.commit()


class Customer(db.Model):
    __tablename__ = "customers"
    id            = db.Column(db.Integer, primary_key=True)
    name          = db.Column(db.String(200), nullable=False)
    email         = db.Column(db.String(200), unique=True, nullable=False, index=True)
    phone         = db.Column(db.String(20), nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    tier          = db.Column(db.String(20), default="regular")  # regular / loyal / vip
    is_active     = db.Column(db.Boolean, default=True)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    orders        = db.relationship("Order", backref="customer", lazy="dynamic",
                                    foreign_keys="Order.customer_id")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def total_orders(self):
        return self.orders.filter(Order.order_status != "cancelled").count()

    @property
    def total_spent(self):
        from sqlalchemy import func
        result = db.session.query(func.sum(Order.total_amount)).filter(
            Order.customer_id == self.id,
            Order.order_status != "cancelled",
        ).scalar()
        return result or 0.0

    def update_tier(self):
        n, s = self.total_orders, self.total_spent
        if n >= 10 or s >= 5000:
            self.tier = "vip"
        elif n >= 3 or s >= 1000:
            self.tier = "loyal"
        else:
            self.tier = "regular"

    def __repr__(self):
        return f"<Customer {self.email}>"


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]


def save_image(file):
    """Save uploaded image and return filename."""
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit(".", 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        return filename
    return None


def allowed_ebook_file(filename):
    return "." in filename and \
           filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EBOOK_EXTENSIONS"]


def save_ebook(file):
    """Save uploaded ebook file and return filename."""
    if file and allowed_ebook_file(file.filename):
        ext = file.filename.rsplit(".", 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.config["EBOOK_FOLDER"], filename))
        return filename
    return None


def save_preview(file):
    """Save uploaded preview PDF and return filename (stored in static/previews/)."""
    if file and file.filename and file.filename.lower().endswith(".pdf"):
        filename = f"preview_{uuid.uuid4().hex}.pdf"
        file.save(os.path.join(app.config["PREVIEW_FOLDER"], filename))
        return filename
    return None


def generate_order_number():
    return "ISKCON" + datetime.now().strftime("%Y%m%d") + uuid.uuid4().hex[:6].upper()


# ── Cart helpers (stored in Flask session) ──

def get_cart():
    return session.get("cart", {})


def save_cart(cart):
    session["cart"] = cart
    session.modified = True


def cart_item_count():
    return sum(item["qty"] for item in get_cart().values())


def cart_totals():
    cart = get_cart()
    if not cart:
        return {"subtotal": 0, "shipping": 0, "discount": 0, "total": 0, "items": []}

    book_ids = [int(k) for k in cart.keys()]
    books = {b.id: b for b in Book.query.filter(Book.id.in_(book_ids)).all()}

    items, subtotal = [], 0
    for book_id_str, item in cart.items():
        book = books.get(int(book_id_str))
        if not book:
            continue
        line_total = book.price * item["qty"]
        subtotal += line_total
        items.append({
            "book":       book,
            "qty":        item["qty"],
            "line_total": line_total,
        })

    shipping = 0 if subtotal >= app.config["FREE_SHIPPING_ABOVE"] else app.config["SHIPPING_CHARGE"]
    discount = session.get("coupon_discount", 0)
    total = max(0, subtotal + shipping - discount)

    return {
        "items":    items,
        "subtotal": subtotal,
        "shipping": shipping,
        "discount": discount,
        "total":    total,
    }


# ── Admin auth ──

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            flash("Please login to access the admin panel.", "warning")
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ── Customer auth ──

def get_current_customer():
    cid = session.get("customer_id")
    if cid:
        return Customer.query.get(cid)
    return None


def customer_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("customer_id"):
            flash("Please log in to access your account.", "warning")
            return redirect(url_for("customer_login", next=request.full_path))
        return f(*args, **kwargs)
    return decorated


# ── Context processors ──

@app.context_processor
def inject_globals():
    return {
        "cart_count":       cart_item_count(),
        "categories":       Category.query.order_by(Category.sort_order).all(),
        "store_name":       app.config["STORE_NAME"],
        "whatsapp_num":     app.config["WHATSAPP_NUMBER"],
        "upi_id":           app.config["UPI_ID"],
        "upi_name":         app.config["UPI_NAME"],
        "current_customer": get_current_customer(),
    }


# ─────────────────────────────────────────────
# SEO — sitemap & robots
# ─────────────────────────────────────────────

@app.route("/api/shipping-rate/<pincode>")
def api_shipping_rate(pincode):
    """AJAX: return Delhivery Prepaid rate for a destination pincode."""
    if not pincode.isdigit() or len(pincode) != 6:
        return jsonify({"serviceable": False, "error": "Invalid pincode"})
    from delhivery import get_shipping_rate, check_serviceability
    svc = check_serviceability(pincode)
    if not svc.get("serviceable"):
        return jsonify({"serviceable": False, "error": "Delhivery does not deliver to this pincode"})
    if not svc.get("pre_paid"):
        return jsonify({"serviceable": False, "error": "Prepaid delivery not available at this pincode"})
    totals = cart_totals()
    weight_grams = max(
        int(sum(item["qty"] * (item["book"].weight_kg or app.config["DELHIVERY_DEFAULT_WEIGHT"]) * 1000
                for item in totals["items"])),
        500
    )
    rate, zone, err = get_shipping_rate(pincode, weight_grams)
    if err or rate is None:
        return jsonify({"serviceable": False, "error": err or "Rate unavailable"})
    return jsonify({
        "serviceable": True,
        "rate":  rate,
        "zone":  zone,
        "city":  svc.get("city", ""),
        "oda":   svc.get("oda", False),
    })


@app.route("/sitemap.xml")
def sitemap():
    domain = "https://iskconbooks.in"
    books  = Book.query.filter_by(active=True, deleted=False).all()
    lines  = ['<?xml version="1.0" encoding="UTF-8"?>',
              '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for path in ["/", "/books", "/order/track"]:
        lines.append(f"<url><loc>{domain}{path}</loc><changefreq>weekly</changefreq><priority>0.8</priority></url>")
    for book in books:
        lines.append(f"<url><loc>{domain}/book/{book.id}</loc><changefreq>monthly</changefreq><priority>0.9</priority></url>")
    lines.append("</urlset>")
    return "\n".join(lines), 200, {"Content-Type": "application/xml"}


@app.route("/robots.txt")
def robots():
    content = """User-agent: *
Allow: /
Disallow: /admin/
Disallow: /checkout
Disallow: /cart
Disallow: /payment/
Disallow: /order/success/
Disallow: /order/failed/

Sitemap: https://iskconbooks.in/sitemap.xml
"""
    return content, 200, {"Content-Type": "text/plain"}


# ─────────────────────────────────────────────
# PUBLIC ROUTES
# ─────────────────────────────────────────────

@app.route("/")
def index():
    featured_books  = Book.query.filter_by(featured=True, active=True).limit(8).all()
    new_arrivals    = Book.query.filter_by(active=True).order_by(Book.created_at.desc()).limit(6).all()
    categories      = Category.query.order_by(Category.sort_order).all()
    # Look up carousel books from admin settings, fall back to keyword search
    def get_carousel_book(setting_key, keyword):
        book_id = Setting.get(setting_key)
        if book_id:
            b = Book.query.filter_by(id=int(book_id), active=True).first()
            if b:
                return b
        return Book.query.filter(Book.title.ilike(f'%{keyword}%'), Book.active == True).order_by(Book.id).first()

    carousel_books = {
        'slide0':  get_carousel_book('carousel_slide_0', 'Prabhupada'),
        'gita':    get_carousel_book('carousel_slide_1', 'Bhagavad Gita'),
        'sb':      get_carousel_book('carousel_slide_2', 'Bhagavatam'),
        'krishna': get_carousel_book('carousel_slide_3', 'Krishna'),
        'nod':     get_carousel_book('carousel_slide_4', 'Nectar of Devotion'),
    }
    return render_template("index.html",
                           featured_books=featured_books,
                           new_arrivals=new_arrivals,
                           categories=categories,
                           carousel_books=carousel_books)


@app.route("/books")
def books():
    query  = request.args.get("q", "").strip()
    cat    = request.args.get("category", "")
    lang   = request.args.get("language", "")
    sort   = request.args.get("sort", "title")
    page   = request.args.get("page", 1, type=int)

    bq = Book.query.filter_by(active=True)

    if query:
        import re
        for word in query.split():
            word = re.sub(r"[^\w]", "", word)
            if not word:
                continue
            bq = bq.filter(or_(
                Book.title.ilike(f"%{word}%"),
                Book.author.ilike(f"%{word}%"),
                Book.description.ilike(f"%{word}%"),
            ))
    if cat:
        category = Category.query.filter_by(slug=cat).first()
        if category:
            bq = bq.filter_by(category_id=category.id)
    if lang:
        bq = bq.filter_by(language=lang)

    sort_map = {
        "title":      Book.title.asc(),
        "price_low":  Book.price.asc(),
        "price_high": Book.price.desc(),
        "newest":     Book.created_at.desc(),
    }
    bq = bq.order_by(sort_map.get(sort, Book.title.asc()))

    pagination   = bq.paginate(page=page, per_page=12, error_out=False)
    languages    = [r[0] for r in db.session.query(Book.language).filter_by(active=True).distinct().all()]
    active_cat   = Category.query.filter_by(slug=cat).first() if cat else None

    return render_template("books.html",
                           books=pagination.items,
                           pagination=pagination,
                           query=query,
                           active_cat=active_cat,
                           language=lang,
                           sort=sort,
                           languages=languages)


@app.route("/book/<int:book_id>")
def book_detail(book_id):
    import re
    book    = Book.query.get_or_404(book_id)
    related = Book.query.filter_by(category_id=book.category_id, active=True)\
                        .filter(Book.id != book_id).limit(4).all()

    # Find same book in other languages by stripping trailing "(Language)" from title
    base_title = re.sub(r'\s*\(\s*[^)]+\)\s*$', '', book.title).strip()
    other_lang_books = Book.query.filter(
        Book.active == True,
        Book.id != book_id,
        Book.title.ilike(f'{base_title}%'),
        Book.language != book.language,
    ).order_by(Book.language).all() if base_title else []

    return render_template("book_detail.html", book=book, related=related,
                           other_lang_books=other_lang_books)


# ─────────────────────────────────────────────
# CART ROUTES
# ─────────────────────────────────────────────

@app.route("/cart/add/<int:book_id>", methods=["POST"])
def add_to_cart(book_id):
    book = Book.query.get_or_404(book_id)
    qty  = int(request.form.get("qty", 1))
    cart = get_cart()
    key  = str(book_id)
    if key in cart:
        cart[key]["qty"] = min(cart[key]["qty"] + qty, book.stock)
    else:
        cart[key] = {"qty": qty, "title": book.title}
    save_cart(cart)
    flash(f'"{book.title}" added to cart!', "success")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"success": True, "cart_count": cart_item_count()})
    return redirect(request.referrer or url_for("cart"))


@app.route("/cart")
def cart():
    totals = cart_totals()
    coupon_code = session.get("coupon_code", "")
    return render_template("cart.html", **totals, coupon_code=coupon_code)


@app.route("/cart/update", methods=["POST"])
def update_cart():
    cart = get_cart()
    for key in list(cart.keys()):
        new_qty = request.form.get(f"qty_{key}", type=int)
        if new_qty is not None:
            if new_qty <= 0:
                del cart[key]
            else:
                cart[key]["qty"] = new_qty
    save_cart(cart)
    # Reset coupon if cart changed
    session.pop("coupon_code", None)
    session.pop("coupon_discount", None)
    flash("Cart updated.", "success")
    return redirect(url_for("cart"))


@app.route("/cart/remove/<int:book_id>")
def remove_from_cart(book_id):
    cart = get_cart()
    cart.pop(str(book_id), None)
    save_cart(cart)
    session.pop("coupon_code", None)
    session.pop("coupon_discount", None)
    flash("Item removed from cart.", "info")
    return redirect(url_for("cart"))


@app.route("/cart/clear")
def clear_cart():
    session.pop("cart", None)
    session.pop("coupon_code", None)
    session.pop("coupon_discount", None)
    return redirect(url_for("cart"))


# ─────────────────────────────────────────────
# COUPON
# ─────────────────────────────────────────────

@app.route("/apply-coupon", methods=["POST"])
def apply_coupon():
    code    = request.form.get("coupon_code", "").strip().upper()
    totals  = cart_totals()
    coupon  = Coupon.query.filter_by(code=code).first()

    if not coupon:
        flash("Invalid coupon code.", "danger")
        return redirect(url_for("cart"))

    valid, msg = coupon.is_valid(totals["subtotal"])
    if not valid:
        flash(msg, "danger")
        return redirect(url_for("cart"))

    discount = coupon.calculate_discount(totals["subtotal"])
    session["coupon_code"]     = code
    session["coupon_discount"] = discount
    session.modified = True
    flash(f"Coupon applied! You saved ₹{discount:.0f}.", "success")
    return redirect(url_for("cart"))


# ─────────────────────────────────────────────
# CHECKOUT & PAYMENT
# ─────────────────────────────────────────────
# PayU helpers
# ─────────────────────────────────────────────

def _payu_hash(key, txnid, amount, productinfo, firstname, email, salt):
    s = f"{key}|{txnid}|{amount}|{productinfo}|{firstname}|{email}|||||||||{salt}"
    return hashlib.sha512(s.encode("utf-8")).hexdigest()

def _payu_verify_hash(data, salt):
    s = (f"{salt}|{data.get('status')}||||||"
         f"{data.get('udf5','')}|{data.get('udf4','')}|{data.get('udf3','')}|"
         f"{data.get('udf2','')}|{data.get('udf1','')}|{data.get('email','')}|"
         f"{data.get('firstname','')}|{data.get('productinfo','')}|"
         f"{data.get('amount','')}|{data.get('txnid','')}|{data.get('key','')}")
    return hashlib.sha512(s.encode("utf-8")).hexdigest()

# ─────────────────────────────────────────────

@app.route("/checkout", methods=["GET", "POST"])
def checkout():
    totals = cart_totals()
    if not totals["items"]:
        flash("Your cart is empty.", "warning")
        return redirect(url_for("books"))

    if request.method == "POST":
        name    = request.form.get("name", "").strip()
        email   = request.form.get("email", "").strip()
        phone   = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        city    = request.form.get("city", "").strip()
        state   = request.form.get("state", "").strip()
        pincode = request.form.get("pincode", "").strip()
        payment = request.form.get("payment_method", "cod")
        app.logger.info(f"[CHECKOUT] payment_method received: '{payment}' | form keys: {list(request.form.keys())}")
        notes   = request.form.get("notes", "").strip()

        if not all([name, phone, address, city, pincode]):
            flash("Please fill all required fields.", "danger")
            return render_template("checkout.html", **totals)

        if payment == "cod":
            flash("Cash on Delivery is not available. Please choose an online payment method.", "danger")
            return render_template("checkout.html", **totals)

        # Calculate Delhivery shipping rate for this pincode
        from delhivery import get_shipping_rate
        weight_grams = max(
            int(sum(item["qty"] * (item["book"].weight_kg or app.config["DELHIVERY_DEFAULT_WEIGHT"]) * 1000
                    for item in totals["items"])),
            500
        )
        delhivery_rate, _, _ = get_shipping_rate(pincode, weight_grams)
        shipping_charge = delhivery_rate if delhivery_rate is not None else totals["shipping"]
        subtotal    = totals["subtotal"]
        discount    = totals["discount"]
        total_amount = max(0, subtotal + shipping_charge - discount)

        # Link to customer account if logged in
        cust_id = session.get("customer_id")

        # Create order
        order = Order(
            order_number    = generate_order_number(),
            customer_name   = name,
            customer_email  = email,
            customer_phone  = phone,
            address         = address,
            city            = city,
            state           = state,
            pincode         = pincode,
            subtotal        = subtotal,
            shipping_charge = shipping_charge,
            discount_amount = discount,
            total_amount    = total_amount,
            payment_method  = payment,
            coupon_code     = session.get("coupon_code"),
            notes           = notes,
            payment_status  = "pending",
            order_status    = "placed",
            customer_id     = cust_id,
        )
        db.session.add(order)
        db.session.flush()  # get order.id

        for cart_item in totals["items"]:
            oi = OrderItem(
                order_id   = order.id,
                book_id    = cart_item["book"].id,
                book_title = cart_item["book"].title,
                quantity   = cart_item["qty"],
                price      = cart_item["book"].price,
            )
            # Reduce stock
            cart_item["book"].stock = max(0, cart_item["book"].stock - cart_item["qty"])
            db.session.add(oi)

        # Update coupon usage
        if order.coupon_code:
            coupon = Coupon.query.filter_by(code=order.coupon_code).first()
            if coupon:
                coupon.used_count += 1

        # Update customer tier if logged in
        if cust_id:
            cust = Customer.query.get(cust_id)
            if cust:
                cust.update_tier()

        db.session.commit()

        # Clear cart & coupon from session
        session.pop("cart", None)
        session.pop("coupon_code", None)
        session.pop("coupon_discount", None)

        if payment == "payu":
            key  = app.config["PAYU_MERCHANT_KEY"]
            salt = app.config["PAYU_MERCHANT_SALT"]
            env  = app.config["PAYU_ENV"]
            amount_str  = f"{order.total_amount:.2f}"
            firstname   = (order.customer_name.split()[0] if order.customer_name else "Customer")[:50]
            email_str   = order.customer_email or ""
            hash_val    = _payu_hash(key, order.order_number, amount_str,
                                     "ISKCON Books", firstname, email_str, salt)
            payu_url    = ("https://test.payu.in/_payment" if env == "test"
                           else "https://secure.payu.in/_payment")
            return render_template("payment_payu.html",
                                   order=order,
                                   payu_url=payu_url,
                                   key=key,
                                   amount=amount_str,
                                   firstname=firstname,
                                   email=email_str,
                                   phone=order.customer_phone or "",
                                   hash_val=hash_val,
                                   surl=url_for("payment_payu_success", _external=True),
                                   furl=url_for("payment_payu_failure", _external=True))

        if payment == "razorpay":
            # Create Razorpay order via direct HTTP (no SDK dependency)
            try:
                import requests as _req
                rp_resp = _req.post(
                    "https://api.razorpay.com/v1/orders",
                    auth=(app.config["RAZORPAY_KEY_ID"], app.config["RAZORPAY_KEY_SECRET"]),
                    json={
                        "amount":   int(order.total_amount * 100),
                        "currency": "INR",
                        "receipt":  order.order_number,
                    },
                    timeout=10,
                )
                rp_resp.raise_for_status()
                rp_order = rp_resp.json()
                order.razorpay_order_id = rp_order["id"]
                db.session.commit()
                return render_template("payment_razorpay.html",
                                       order=order,
                                       rp_order=rp_order,
                                       key_id=app.config["RAZORPAY_KEY_ID"])
            except Exception as e:
                # Delete child rows first to satisfy FK constraint (PostgreSQL)
                OrderItem.query.filter_by(order_id=order.id).delete()
                db.session.delete(order)
                db.session.commit()
                # Restore cart and stock
                restored_cart = {}
                for item in totals["items"]:
                    key = str(item["book"].id)
                    restored_cart[key] = {"qty": item["qty"], "title": item["book"].title}
                    item["book"].stock += item["qty"]
                db.session.commit()
                session["cart"] = restored_cart
                session.modified = True
                flash(f"Payment gateway error: {e}. Please try again or choose Cash on Delivery.", "danger")
                return redirect(url_for("checkout"))

        if payment == "upi":
            return redirect(url_for("payment_upi_qr", order_number=order.order_number))

        flash(f"Order #{order.order_number} placed successfully! 🎉", "success")
        return redirect(url_for("order_success", order_number=order.order_number))

    cust = get_current_customer()
    return render_template("checkout.html", **totals,
                           razorpay_key=app.config["RAZORPAY_KEY_ID"],
                           prefill=cust)


@app.route("/payment/verify", methods=["POST"])
def payment_verify():
    data = request.get_json() or request.form.to_dict()
    order_number = data.get("order_number") or data.get("receipt")
    order = Order.query.filter_by(order_number=order_number).first()
    if not order:
        return jsonify({"success": False, "error": "Order not found"}), 404

    # Verify Razorpay signature
    rp_order_id   = data.get("razorpay_order_id")
    rp_payment_id = data.get("razorpay_payment_id")
    rp_signature  = data.get("razorpay_signature")

    try:
        msg = f"{rp_order_id}|{rp_payment_id}".encode()
        expected = hmac.new(app.config["RAZORPAY_KEY_SECRET"].encode(), msg, hashlib.sha256).hexdigest()
        if hmac.compare_digest(expected, rp_signature):
            order.payment_status    = "paid"
            order.razorpay_payment_id = rp_payment_id
            order.order_status      = "confirmed"
            db.session.commit()
            send_order_confirmation(order)
            return jsonify({"success": True, "redirect": url_for("order_success", order_number=order.order_number)})
    except Exception:
        pass

    order.payment_status = "failed"
    db.session.commit()
    return jsonify({"success": False, "redirect": url_for("payment_failed", order_number=order.order_number)})


@app.route("/order/success/<order_number>")
def order_success(order_number):
    order = Order.query.filter_by(order_number=order_number).first_or_404()
    return render_template("payment_success.html", order=order)


@app.route("/payment/payu/success", methods=["POST"])
def payment_payu_success():
    data = request.form.to_dict()
    salt = app.config["PAYU_MERCHANT_SALT"]
    txnid = data.get("txnid", "")
    order = Order.query.filter_by(order_number=txnid).first()
    if order:
        expected = _payu_verify_hash(data, salt)
        received = data.get("hash", "")
        if expected == received and data.get("status") == "success":
            order.payment_status = "paid"
            order.order_status = "confirmed"
            order.razorpay_payment_id = data.get("mihpayid", "")
            db.session.commit()
            send_order_confirmation(order)
            flash("Payment successful! 🎉 Hare Krishna!", "success")
            return redirect(url_for("order_success", order_number=order.order_number))
        else:
            order.payment_status = "failed"
            db.session.commit()
    flash("Payment verification failed. Please contact support.", "danger")
    return redirect(url_for("order_track"))


@app.route("/payment/payu/failure", methods=["POST"])
def payment_payu_failure():
    data = request.form.to_dict()
    txnid = data.get("txnid", "")
    order = Order.query.filter_by(order_number=txnid).first()
    if order:
        order.payment_status = "failed"
        db.session.commit()
    flash("Payment failed or cancelled. Please try again.", "danger")
    return redirect(url_for("order_track"))


@app.route("/payment/payu/webhook", methods=["POST"])
def payment_payu_webhook():
    """PayU S2S server-to-server webhook — fires independently of browser redirect."""
    data = request.form.to_dict()
    salt = app.config["PAYU_MERCHANT_SALT"]
    txnid = data.get("txnid", "")
    order = Order.query.filter_by(order_number=txnid).first()
    if order:
        expected = _payu_verify_hash(data, salt)
        received = data.get("hash", "")
        if expected == received:
            newly_confirmed = False
            if data.get("status") == "success" and order.payment_status != "paid":
                order.payment_status = "paid"
                order.order_status = "confirmed"
                order.razorpay_payment_id = data.get("mihpayid", "")
                newly_confirmed = True
            elif data.get("status") != "success" and order.payment_status == "pending":
                order.payment_status = "failed"
            db.session.commit()
            if newly_confirmed:
                send_order_confirmation(order)
    return "OK", 200


@app.route("/payment/upi-qr/<order_number>")
def payment_upi_qr(order_number):
    order = Order.query.filter_by(order_number=order_number).first_or_404()
    upi_id   = app.config.get("UPI_ID", "")
    upi_name = app.config.get("UPI_NAME", "ISKCON Book Store")
    amount   = f"{order.total_amount:.2f}"
    upi_link = (
        f"upi://pay?pa={upi_id}&pn={upi_name.replace(' ', '%20')}"
        f"&am={amount}&tn={order.order_number}&cu=INR"
    )
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M,
                           box_size=8, border=4)
        qr.add_data(upi_link)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#1a1a1a", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_b64 = base64.b64encode(buf.getvalue()).decode()
    except Exception:
        qr_b64 = None
    return render_template("payment_upi_qr.html",
                           order=order, qr_b64=qr_b64,
                           upi_id=upi_id, upi_name=upi_name, amount=amount)


@app.route("/order/upi-confirm/<order_number>", methods=["POST"])
def upi_confirm(order_number):
    order = Order.query.filter_by(order_number=order_number).first_or_404()
    if order.payment_method != "upi":
        abort(400)
    utr = request.form.get("utr", "").strip()
    if utr:
        try:
            # Use raw SQL in case column was just added and SQLAlchemy mapper cache is stale
            db.session.execute(
                db.text("UPDATE orders SET upi_transaction_id = :utr WHERE id = :oid"),
                {"utr": utr, "oid": order.id}
            )
            db.session.commit()
            flash("Transaction ID submitted! We will verify and confirm your order shortly.", "success")
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] upi_confirm save failed: {e}")
            flash("Could not save Transaction ID. Please send it via WhatsApp.", "warning")
    return redirect(url_for("order_success", order_number=order_number))


@app.route("/ebook/download/<order_number>/<int:book_id>")
def ebook_download(order_number, book_id):
    order = Order.query.filter_by(order_number=order_number).first_or_404()

    # Access control: COD denied; Razorpay must be paid; UPI allowed (trust-based)
    if order.payment_method == "cod":
        abort(403)
    if order.payment_method == "razorpay" and order.payment_status != "paid":
        abort(403)

    # Verify book is actually in the order
    item = next((i for i in order.items if i.book_id == book_id), None)
    if not item:
        abort(404)

    book = Book.query.get_or_404(book_id)
    if not book.is_ebook or not book.ebook_file:
        abort(404)

    ebook_path = os.path.join(app.config["EBOOK_FOLDER"], book.ebook_file)
    if not os.path.exists(ebook_path):
        flash("eBook file not found. Please contact support.", "warning")
        return redirect(url_for("order_success", order_number=order_number))

    ext = book.ebook_file.rsplit(".", 1)[1]
    return send_from_directory(
        app.config["EBOOK_FOLDER"],
        book.ebook_file,
        as_attachment=True,
        download_name=f"{book.title}.{ext}"
    )


@app.route("/order/failed/<order_number>")
def payment_failed(order_number):
    order = Order.query.filter_by(order_number=order_number).first_or_404()
    return render_template("payment_failed.html", order=order)


@app.route("/order/track", methods=["GET", "POST"])
def order_track():
    order = None
    if request.method == "POST":
        query = request.form.get("query", "").strip()
        # Normalise phone: try both raw input and without +91 / 91 prefix
        phone_variants = [query]
        digits = query.lstrip("+")
        if digits.startswith("91") and len(digits) == 12:
            phone_variants.append(digits[2:])       # strip 91 → 10-digit
        elif len(digits) == 10:
            phone_variants.append("91" + digits)    # add 91
            phone_variants.append("+91" + digits)   # add +91
        order = Order.query.filter(
            or_(Order.order_number == query,
                Order.customer_phone.in_(phone_variants))
        ).order_by(Order.created_at.desc()).first()
        if not order:
            flash("No order found with that order number or phone.", "warning")
    return render_template("order_tracking.html", order=order)


# ─────────────────────────────────────────────
# CUSTOMER ACCOUNT ROUTES
# ─────────────────────────────────────────────

@app.route("/account/register", methods=["GET", "POST"])
def customer_register():
    if session.get("customer_id"):
        return redirect(url_for("customer_dashboard"))

    if request.method == "POST":
        name     = request.form.get("name", "").strip()
        email    = request.form.get("email", "").strip().lower()
        phone    = request.form.get("phone", "").strip()
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm_password", "")

        error = None
        if not all([name, email, phone, password]):
            error = "All fields are required."
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        elif password != confirm:
            error = "Passwords do not match."
        elif Customer.query.filter_by(email=email).first():
            error = "An account with this email already exists."

        if error:
            flash(error, "danger")
            return render_template("account/register.html",
                                   name=name, email=email, phone=phone)

        customer = Customer(name=name, email=email, phone=phone)
        customer.set_password(password)
        db.session.add(customer)
        db.session.flush()  # get customer.id

        # Auto-link existing orders placed with same email or phone
        linked = Order.query.filter(
            Order.customer_id.is_(None),
            or_(Order.customer_email == email, Order.customer_phone == phone),
        ).all()
        for o in linked:
            o.customer_id = customer.id
        db.session.flush()
        customer.update_tier()
        db.session.commit()

        session["customer_id"] = customer.id
        flash(f"Welcome, {customer.name}! Your account has been created.", "success")
        if linked:
            flash(f"{len(linked)} previous order(s) have been linked to your account.", "info")
        return redirect(url_for("customer_dashboard"))

    return render_template("account/register.html",
                           name=request.args.get("name", ""),
                           email=request.args.get("email", ""),
                           phone=request.args.get("phone", ""))


@app.route("/account/login", methods=["GET", "POST"])
def customer_login():
    if session.get("customer_id"):
        return redirect(url_for("customer_dashboard"))

    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        customer = Customer.query.filter_by(email=email, is_active=True).first()

        if customer and customer.check_password(password):
            session["customer_id"] = customer.id
            raw_next = request.form.get("next", "")
            # Only allow local redirects — reject absolute URLs to prevent open redirect
            from urllib.parse import urlparse
            next_url = raw_next if raw_next and not urlparse(raw_next).netloc else url_for("customer_dashboard")
            flash(f"Welcome back, {customer.name}!", "success")
            return redirect(next_url)

        flash("Invalid email or password.", "danger")
        return render_template("account/login.html", email=email,
                               next=request.form.get("next", ""))

    return render_template("account/login.html", email="",
                           next=request.args.get("next", ""))


@app.route("/account/logout")
def customer_logout():
    session.pop("customer_id", None)
    flash("You have been logged out.", "info")
    return redirect(url_for("index"))


@app.route("/account/")
@customer_login_required
def customer_dashboard():
    customer = get_current_customer()
    orders = customer.orders.order_by(Order.created_at.desc()).all()
    return render_template("account/dashboard.html", customer=customer, orders=orders)


@app.route("/account/profile", methods=["GET", "POST"])
@customer_login_required
def customer_profile():
    customer = get_current_customer()

    if request.method == "POST":
        name         = request.form.get("name", "").strip()
        phone        = request.form.get("phone", "").strip()
        new_password = request.form.get("new_password", "")
        confirm      = request.form.get("confirm_password", "")

        if not name or not phone:
            flash("Name and phone are required.", "danger")
        elif new_password and len(new_password) < 8:
            flash("New password must be at least 8 characters.", "danger")
        elif new_password and new_password != confirm:
            flash("Passwords do not match.", "danger")
        else:
            customer.name  = name
            customer.phone = phone
            if new_password:
                customer.set_password(new_password)
            db.session.commit()
            flash("Profile updated successfully.", "success")
            return redirect(url_for("customer_profile"))

    return render_template("account/profile.html", customer=customer)


# ─────────────────────────────────────────────
# ADMIN ROUTES
# ─────────────────────────────────────────────

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if session.get("admin_logged_in"):
        return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if (username == app.config["ADMIN_USERNAME"] and
                check_password_hash(app.config["ADMIN_PASSWORD_HASH"], password)):
            session["admin_logged_in"] = True
            session.permanent = True
            flash("Welcome back, Admin! 🙏", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Invalid credentials.", "danger")
    return render_template("admin/login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    flash("Logged out.", "info")
    return redirect(url_for("admin_login"))


@app.route("/admin/test-email", methods=["GET", "POST"])
@admin_required
def admin_test_email():
    """Send a test email via Brevo HTTP API to verify config is working."""
    import requests as http_req
    result = None
    cfg = {
        "api_key_set": bool(app.config.get("BREVO_API_KEY", "")),
        "username":    app.config.get("MAIL_USERNAME", "iskconbooks.in@gmail.com"),
    }
    if request.method == "POST":
        to = request.form.get("to_email", "").strip()
        if not to:
            result = {"ok": False, "msg": "Please enter a recipient email address."}
        elif not cfg["api_key_set"]:
            result = {"ok": False, "msg": "BREVO_API_KEY not set in Railway environment variables."}
        else:
            payload = {
                "sender":      {"name": "ISKCON Book Store", "email": cfg["username"]},
                "to":          [{"email": to}],
                "subject":     "Test Email — ISKCON Book Store",
                "htmlContent": "<h2>Test Email</h2><p>If you received this, Brevo email is configured correctly. Hare Krishna! 🙏</p>",
            }
            headers = {
                "accept":       "application/json",
                "content-type": "application/json",
                "api-key":      app.config.get("BREVO_API_KEY"),
            }
            try:
                resp = http_req.post("https://api.brevo.com/v3/smtp/email",
                                     json=payload, headers=headers, timeout=15)
                if resp.status_code in (200, 201):
                    result = {"ok": True, "msg": f"Test email sent to {to}. Check inbox and spam folder."}
                else:
                    result = {"ok": False, "msg": f"Brevo API error {resp.status_code}: {resp.text[:300]}"}
            except Exception as e:
                result = {"ok": False, "msg": f"Request error: {e}"}
    to_email = request.form.get("to_email", cfg["username"]) if request.method == "POST" else cfg["username"]
    return render_template("admin/test_email.html", cfg=cfg, result=result, to_email=to_email)


@app.route("/admin/")
@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    total_orders        = Order.query.filter_by(is_deleted=False).count()
    total_revenue       = db.session.query(db.func.sum(Order.total_amount))\
                                    .filter(Order.payment_status != "failed",
                                            Order.is_deleted == False).scalar() or 0
    total_books         = Book.query.filter_by(active=True).count()
    pending_orders      = Order.query.filter_by(order_status="placed", is_deleted=False).count()
    recent_orders       = Order.query.filter_by(is_deleted=False)\
                                     .order_by(Order.created_at.desc()).limit(10).all()
    low_stock           = Book.query.filter(Book.stock < 5, Book.active == True).all()
    temple_books_total  = db.session.query(db.func.sum(StockReceipt.quantity)).scalar() or 0
    temple_pending_payment = db.session.query(db.func.sum(StockReceipt.total_payment))\
                               .filter(StockReceipt.payment_status == "pending").scalar() or 0
    recent_receipts     = StockReceipt.query.order_by(StockReceipt.received_date.desc()).limit(5).all()
    return render_template("admin/dashboard.html",
                           total_orders=total_orders,
                           total_revenue=total_revenue,
                           total_books=total_books,
                           pending_orders=pending_orders,
                           recent_orders=recent_orders,
                           low_stock=low_stock,
                           temple_books_total=temple_books_total,
                           temple_pending_payment=temple_pending_payment,
                           recent_receipts=recent_receipts)


# ── Admin: Books ──

@app.route("/admin/books")
@admin_required
def admin_books():
    page          = request.args.get("page", 1, type=int)
    query         = request.args.get("q", "")
    category_id   = request.args.get("category_id", type=int)
    title_f       = request.args.get("tf", "")
    author_f      = request.args.get("af", "")
    lang_filter   = request.args.get("lang", "")
    fmt_filter    = request.args.get("fmt", "")
    status_filter = request.args.get("status", "")

    bq = Book.query.filter_by(deleted=False)
    if query:
        import re
        for word in query.split():
            word = re.sub(r"[^\w]", "", word)
            if not word:
                continue
            bq = bq.filter(or_(
                Book.title.ilike(f"%{word}%"),
                Book.author.ilike(f"%{word}%"),
            ))
    if title_f:
        bq = bq.filter(Book.title.ilike(f"%{title_f}%"))
    if author_f:
        bq = bq.filter(Book.author.ilike(f"%{author_f}%"))
    if category_id:
        bq = bq.filter_by(category_id=category_id)
    if lang_filter:
        bq = bq.filter(Book.language == lang_filter)
    if fmt_filter == 'ebook':
        bq = bq.filter(Book.is_ebook == True)
    elif fmt_filter == 'paper':
        bq = bq.filter(Book.is_ebook == False)
    if status_filter == 'active':
        bq = bq.filter(Book.active == True)
    elif status_filter == 'inactive':
        bq = bq.filter(Book.active == False)

    bq = bq.order_by(Book.created_at.desc())

    books          = bq.paginate(page=page, per_page=20)
    trash_count    = Book.query.filter_by(deleted=True).count()
    all_categories = Category.query.order_by(Category.sort_order).all()
    active_cat     = Category.query.get(category_id) if category_id else None
    langs          = [r[0] for r in db.session.query(Book.language)
                      .filter(Book.deleted == False, Book.language != None, Book.language != '')
                      .distinct().order_by(Book.language).all()]
    return render_template("admin/books.html", books=books, query=query,
                           trash_count=trash_count, all_categories=all_categories,
                           active_cat=active_cat, category_id=category_id,
                           title_f=title_f, author_f=author_f,
                           lang_filter=lang_filter, fmt_filter=fmt_filter,
                           status_filter=status_filter, langs=langs)


@app.route("/admin/books/export-stock-csv")
@admin_required
def export_stock_csv():
    """Download all books as a CSV stock receipt for physical records."""
    all_books = Book.query.order_by(Book.category_id, Book.title).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Sr No", "Title", "Author", "Category", "Language", "Format",
        "Price (INR)", "Original Price (INR)", "Discount %",
        "Stock Qty", "ISBN", "Publisher", "Pages", "Active", "Featured"
    ])

    for idx, book in enumerate(all_books, start=1):
        writer.writerow([
            idx,
            book.title,
            book.author,
            book.category.name if book.category else "",
            book.language or "",
            "eBook" if book.is_ebook else "Paper",
            int(book.price),
            int(book.original_price) if book.original_price else "",
            book.discount_percent or "",
            book.stock,
            book.isbn or "",
            book.publisher or "",
            book.pages or "",
            "Yes" if book.active else "No",
            "Yes" if book.featured else "No",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")   # utf-8-sig adds BOM for Excel
    from flask import Response
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/admin/books/add", methods=["GET", "POST"])
@admin_required
def admin_add_book():
    categories = Category.query.order_by(Category.name).all()
    if request.method == "POST":
        image_file = request.files.get("image")
        image_name = save_image(image_file) or "default_book.jpg"

        ebook_file_obj = request.files.get("ebook_file")
        is_ebook = request.form.get("book_format") == "ebook"
        ebook_filename = save_ebook(ebook_file_obj) if is_ebook else None
        preview_filename = save_preview(request.files.get("preview_file"))

        book = Book(
            title          = request.form["title"].strip(),
            author         = request.form["author"].strip(),
            description    = request.form.get("description", "").strip(),
            short_desc     = request.form.get("short_desc", "").strip(),
            price          = float(request.form["price"]),
            original_price = float(request.form["original_price"]) if request.form.get("original_price") else None,
            image          = image_name,
            category_id    = int(request.form.get("category_id") or 0) if request.form.get("category_id", "") != "" else None,
            isbn           = request.form.get("isbn", "").strip(),
            language       = request.form.get("language", "English").strip(),
            pages          = int(request.form["pages"]) if request.form.get("pages") else None,
            weight_kg      = float(request.form["weight_kg"]) if request.form.get("weight_kg") else 0.1,
            publisher      = request.form.get("publisher", "The Bhaktivedanta Book Trust").strip(),
            stock          = int(request.form.get("stock", 100)),
            featured       = bool(request.form.get("featured")),
            active         = bool(request.form.get("active", True)),
            is_ebook       = is_ebook,
            ebook_file     = ebook_filename,
            preview_file   = preview_filename,
        )
        db.session.add(book)
        db.session.commit()
        flash("Book added successfully!", "success")
        return redirect(url_for("admin_books"))

    return render_template("admin/book_form.html", book=None, categories=categories)


@app.route("/admin/books/edit/<int:book_id>", methods=["GET", "POST"])
@admin_required
def admin_edit_book(book_id):
    book       = Book.query.get_or_404(book_id)
    categories = Category.query.order_by(Category.name).all()

    if request.method == "POST":
        image_file = request.files.get("image")
        if image_file and image_file.filename:
            # Delete old image if not default
            if book.image != "default_book.jpg":
                old_path = os.path.join(app.config["UPLOAD_FOLDER"], book.image)
                if os.path.exists(old_path):
                    os.remove(old_path)
            book.image = save_image(image_file) or book.image

        book.title          = request.form["title"].strip()
        book.author         = request.form["author"].strip()
        book.description    = request.form.get("description", "").strip()
        book.short_desc     = request.form.get("short_desc", "").strip()
        book.price          = float(request.form["price"])
        book.original_price = float(request.form["original_price"]) if request.form.get("original_price") else None
        book.category_id    = int(request.form["category_id"]) if request.form.get("category_id") else None
        book.isbn           = request.form.get("isbn", "").strip()
        book.language       = request.form.get("language", "English").strip()
        book.pages          = int(request.form["pages"]) if request.form.get("pages") else None
        book.weight_kg      = float(request.form["weight_kg"]) if request.form.get("weight_kg") else 0.1
        book.publisher      = request.form.get("publisher", "").strip()
        book.stock          = int(request.form.get("stock", 100))
        book.featured       = bool(request.form.get("featured"))
        book.active         = bool(request.form.get("active"))

        is_ebook = request.form.get("book_format") == "ebook"
        book.is_ebook = is_ebook

        ebook_file_obj = request.files.get("ebook_file")
        if ebook_file_obj and ebook_file_obj.filename:
            if book.ebook_file:
                old_ebook = os.path.join(app.config["EBOOK_FOLDER"], book.ebook_file)
                if os.path.exists(old_ebook):
                    os.remove(old_ebook)
            book.ebook_file = save_ebook(ebook_file_obj)

        if not is_ebook:
            book.ebook_file = None

        preview_file_obj = request.files.get("preview_file")
        if preview_file_obj and preview_file_obj.filename:
            if book.preview_file:
                old_preview = os.path.join(app.config["PREVIEW_FOLDER"], book.preview_file)
                if os.path.exists(old_preview):
                    os.remove(old_preview)
            book.preview_file = save_preview(preview_file_obj)

        db.session.commit()
        flash("Book updated!", "success")
        page = request.form.get("page", 1, type=int)
        return redirect(url_for("admin_books", page=page))

    return render_template("admin/book_form.html", book=book, categories=categories)


@app.route("/admin/books/delete/<int:book_id>", methods=["POST"])
@admin_required
def admin_delete_book(book_id):
    book = Book.query.get_or_404(book_id)
    book.active  = False
    book.deleted = True   # Move to Trash
    db.session.commit()
    flash(f'Book "{book.title}" moved to Trash. You can restore or permanently delete it from the Trash tab.', "info")
    return redirect(url_for("admin_books"))


@app.route("/admin/books/quick-update/<int:book_id>", methods=["POST"])
@admin_required
def admin_quick_update_book(book_id):
    book  = Book.query.get_or_404(book_id)
    price = request.form.get("price", type=float)
    stock = request.form.get("stock", type=int)
    if price is not None and price >= 0:
        book.price = price
    if stock is not None and stock >= 0:
        book.stock = stock
    db.session.commit()
    return jsonify({"success": True})


@app.route("/admin/books/reset-weights", methods=["POST"])
@admin_required
def admin_reset_book_weights():
    """Bulk reset all books with weight=0.5 (old default) to 0.1 kg."""
    updated = Book.query.filter(
        Book.deleted == False,
        Book.weight_kg == 0.5
    ).update({"weight_kg": 0.1}, synchronize_session=False)
    db.session.commit()
    flash(f"Done! {updated} book(s) reset from 0.5 kg to 0.1 kg. Update individual books with correct weights as needed.", "success")
    return redirect(url_for("admin_books"))


@app.route("/admin/books/trash")
@admin_required
def admin_trash_books():
    page          = request.args.get("page", 1, type=int)
    deleted_books = Book.query.filter_by(deleted=True).order_by(Book.created_at.desc()).paginate(page=page, per_page=20)
    return render_template("admin/trash_books.html", books=deleted_books)


@app.route("/admin/books/restore/<int:book_id>", methods=["POST"])
@admin_required
def admin_restore_book(book_id):
    book = Book.query.get_or_404(book_id)
    book.deleted = False
    book.active  = True
    db.session.commit()
    flash(f'Book "{book.title}" restored. You can now edit it.', "success")
    return redirect(url_for("admin_trash_books"))


@app.route("/admin/books/hard-delete/<int:book_id>", methods=["POST"])
@admin_required
def admin_hard_delete_book(book_id):
    book = Book.query.get_or_404(book_id)
    title = book.title
    db.session.delete(book)
    db.session.commit()
    flash(f'Book "{title}" permanently deleted.', "danger")
    return redirect(url_for("admin_trash_books"))




@app.route("/admin/books/toggle-featured/<int:book_id>", methods=["POST"])
@admin_required
def toggle_featured(book_id):
    book = Book.query.get_or_404(book_id)
    book.featured = not book.featured
    db.session.commit()
    return jsonify({"featured": book.featured})


# ── Admin: Temple Stock ──

@app.route("/admin/stock")
@admin_required
def admin_stock():
    page     = request.args.get("page", 1, type=int)
    receipts = StockReceipt.query.order_by(StockReceipt.received_date.desc()).paginate(page=page, per_page=20)
    books    = Book.query.order_by(Book.title).all()
    total_books_received = db.session.query(db.func.sum(StockReceipt.quantity)).scalar() or 0
    total_paid           = db.session.query(db.func.sum(StockReceipt.total_payment))\
                              .filter(StockReceipt.payment_status == "paid").scalar() or 0
    total_pending        = db.session.query(db.func.sum(StockReceipt.total_payment))\
                              .filter(StockReceipt.payment_status == "pending").scalar() or 0
    return render_template("admin/stock.html",
                           receipts=receipts,
                           books=books,
                           total_books_received=total_books_received,
                           total_paid=total_paid,
                           total_pending=total_pending,
                           now=datetime.utcnow())


@app.route("/admin/stock/export-csv")
@admin_required
def export_stock_receipts_csv():
    """Download temple stock receipts as CSV — all / paid / pending."""
    from flask import Response

    payment_filter = request.args.get("payment_status", "")

    rq = StockReceipt.query
    if payment_filter:
        rq = rq.filter_by(payment_status=payment_filter)
    all_receipts = rq.order_by(StockReceipt.received_date.desc()).all()

    # Totals for summary rows
    total_qty     = sum(r.quantity for r in all_receipts)
    total_cost    = sum(r.total_payment for r in all_receipts)

    output = io.StringIO()
    writer = csv.writer(output)

    # Title block
    label = {"paid": "Payment Done", "pending": "Payment Pending"}.get(payment_filter, "All Receipts")
    writer.writerow([f"ISKCON Book Store — Temple Stock Receipt Report ({label})"])
    writer.writerow([f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}"])
    writer.writerow([])

    # Header
    writer.writerow([
        "Sr No", "Date Received", "Book Title",
        "Qty Received", "Cost / Copy (INR)", "Total Amount (INR)",
        "Payment Status", "Notes"
    ])

    for idx, r in enumerate(all_receipts, start=1):
        writer.writerow([
            idx,
            r.received_date.strftime("%d-%m-%Y"),
            r.book_name,
            r.quantity,
            int(r.cost_per_unit),
            int(r.total_payment),
            "PAID" if r.payment_status == "paid" else "PENDING",
            r.notes or "",
        ])

    # Summary footer
    writer.writerow([])
    writer.writerow(["", "", "TOTAL", total_qty, "", int(total_cost), "", ""])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    tag = f"_{payment_filter}" if payment_filter else ""
    filename = f"temple_stock{tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/admin/stock/download-template")
@admin_required
def stock_download_template():
    """Generate an XLSX upload template with book-name dropdowns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from flask import Response

    books = Book.query.filter_by(active=True, deleted=False).order_by(Book.title).all()
    book_titles = [b.title for b in books]

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Upload"

    # Hidden sheet for book names (avoids the 255-char formula limit for long lists)
    ws_books = wb.create_sheet("BookList")
    ws_books.sheet_state = "hidden"
    for i, title in enumerate(book_titles, start=1):
        ws_books.cell(row=i, column=1, value=title)

    orange_fill = PatternFill("solid", fgColor="FF9933")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, bottom=thin, top=thin)

    headers = [
        "Date Received (dd-mm-yyyy)",
        "Book Name",
        "Quantity",
        "Cost/Copy (₹)",
        "Payment Status (Paid / Pending)",
        "Notes",
    ]
    col_widths = [28, 48, 12, 18, 34, 40]

    ws.row_dimensions[1].height = 30
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = orange_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sample row so user knows the format
    sample = [
        datetime.now().strftime("%d-%m-%Y"),
        book_titles[0] if book_titles else "Bhagavad Gita As It Is",
        25,
        80,
        "Paid",
        "Received from temple counter",
    ]
    sample_font = Font(italic=True, color="888888")
    for col, val in enumerate(sample, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = sample_font
        cell.border = border

    # Book name dropdown (uses hidden BookList sheet — no char limit)
    if book_titles:
        n = len(book_titles)
        dv_book = DataValidation(
            type="list",
            formula1=f"BookList!$A$1:$A${n}",
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Book",
            error="Please select a book name from the dropdown.",
        )
        ws.add_data_validation(dv_book)
        dv_book.sqref = "B2:B500"

    # Payment status dropdown
    dv_pay = DataValidation(
        type="list",
        formula1='"Paid,Pending"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Enter Paid or Pending.",
    )
    ws.add_data_validation(dv_pay)
    dv_pay.sqref = "E2:E500"

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"temple_stock_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        out.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/stock/upload", methods=["POST"])
@admin_required
def stock_bulk_upload():
    """Bulk-insert temple stock receipts from an uploaded XLSX or CSV file."""
    uploaded = request.files.get("stock_file")
    if not uploaded or not uploaded.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("admin_stock"))

    ext = uploaded.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "csv"):
        flash("Only .xlsx or .csv files are accepted.", "danger")
        return redirect(url_for("admin_stock"))

    all_books = Book.query.all()
    book_map = {b.title.strip().lower(): b for b in all_books}

    batch_ref = "B-" + datetime.now().strftime("%d%m%y") + "-" + uuid.uuid4().hex[:4].upper()

    raw_rows = []
    if ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c for c in row if c not in (None, "")):
                continue
            raw_rows.append(list(row))
    else:
        content = uploaded.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        next(reader, None)  # skip header
        for r in reader:
            if not any(c.strip() for c in r):
                continue
            raw_rows.append(r)

    inserted = 0
    skipped = []

    for i, row in enumerate(raw_rows, start=2):
        try:
            # Pad short rows
            while len(row) < 6:
                row.append("")

            date_raw    = str(row[0] or "").strip()
            book_name   = str(row[1] or "").strip()
            quantity    = int(float(str(row[2] or 0).strip() or 0))
            cost        = float(str(row[3] or 0).strip() or 0)
            payment_raw = str(row[4] or "paid").strip().lower()
            notes       = str(row[5] or "").strip()

            if not book_name:
                skipped.append(f"Row {i}: no book name")
                continue
            if quantity <= 0:
                skipped.append(f"Row {i}: quantity must be > 0")
                continue

            received_date = datetime.utcnow()
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    received_date = datetime.strptime(date_raw, fmt)
                    break
                except ValueError:
                    pass

            payment_status = "paid" if "paid" in payment_raw else "pending"
            book = book_map.get(book_name.lower())

            if not book:
                skipped.append(f"Row {i}: \"{book_name}\" not found in store — add it via Books first")
                continue

            receipt = StockReceipt(
                book_id       = book.id,
                book_name     = book.title,
                quantity      = quantity,
                cost_per_unit = cost,
                total_payment = round(quantity * cost, 2),
                payment_status= payment_status,
                received_date = received_date,
                notes         = notes,
                batch_ref     = batch_ref,
            )
            db.session.add(receipt)
            book.stock += quantity
            inserted += 1
        except Exception as e:
            skipped.append(f"Row {i}: {e}")

    db.session.commit()

    if inserted:
        flash(f"Bulk upload complete: {inserted} receipt(s) imported successfully.", "success")
    if skipped:
        details = "; ".join(skipped[:5]) + (" …" if len(skipped) > 5 else "")
        flash(
            f"{len(skipped)} row(s) skipped — {details}. "
            "Add missing books via the Books section, then re-upload.",
            "warning",
        )
    if not inserted and not skipped:
        flash("The file had no data rows to import.", "warning")

    return redirect(url_for("admin_stock"))


@app.route("/admin/stock/add", methods=["POST"])
@admin_required
def admin_add_stock():
    book_id      = request.form.get("book_id")
    book         = Book.query.get(book_id) if book_id else None
    quantity     = int(request.form["quantity"])
    cost_per_unit = float(request.form["cost_per_unit"])
    received_date_str = request.form.get("received_date", "")
    received_date = datetime.strptime(received_date_str, "%Y-%m-%d") if received_date_str else datetime.utcnow()

    receipt = StockReceipt(
        book_id        = book.id if book else None,
        book_name      = book.title if book else request.form.get("book_name_manual", "Unknown"),
        quantity       = quantity,
        cost_per_unit  = cost_per_unit,
        total_payment  = round(quantity * cost_per_unit, 2),
        payment_status = request.form.get("payment_status", "paid"),
        received_date  = received_date,
        notes          = request.form.get("notes", "").strip(),
    )
    db.session.add(receipt)
    # Also update book stock
    if book:
        book.stock += quantity
    db.session.commit()
    flash(f"Stock receipt added: {receipt.quantity} copies of '{receipt.book_name}'.", "success")
    return redirect(url_for("admin_stock"))


@app.route("/admin/stock/delete/<int:receipt_id>", methods=["POST"])
@admin_required
def admin_delete_stock(receipt_id):
    receipt = StockReceipt.query.get_or_404(receipt_id)
    # Reverse the stock addition
    if receipt.book_id:
        book = Book.query.get(receipt.book_id)
        if book:
            book.stock = max(0, book.stock - receipt.quantity)
    db.session.delete(receipt)
    db.session.commit()
    flash("Stock receipt deleted.", "info")
    return redirect(url_for("admin_stock"))


# ── Admin: Categories ──

@app.route("/admin/categories")
@admin_required
def admin_categories():
    categories = Category.query.order_by(Category.sort_order).all()
    return render_template("admin/categories.html", categories=categories)


@app.route("/admin/categories/add", methods=["POST"])
@admin_required
def admin_add_category():
    name = request.form["name"].strip()
    slug = name.lower().replace(" ", "-").replace("'", "")
    if not Category.query.filter_by(slug=slug).first():
        cat = Category(
            name        = name,
            slug        = slug,
            description = request.form.get("description", ""),
            icon        = request.form.get("icon", "B"),
            sort_order  = int(request.form.get("sort_order", 0)),
        )
        db.session.add(cat)
        db.session.commit()
        flash(f'Category "{name}" added.', "success")
    else:
        flash("Category already exists.", "warning")
    return redirect(url_for("admin_categories"))


@app.route("/admin/categories/delete/<int:cat_id>", methods=["POST"])
@admin_required
def admin_delete_category(cat_id):
    cat = Category.query.get_or_404(cat_id)
    db.session.delete(cat)
    db.session.commit()
    flash(f'Category "{cat.name}" deleted.', "info")
    return redirect(url_for("admin_categories"))


# ── Admin: Orders ──

@app.route("/admin/orders")
@admin_required
def admin_orders():
    page           = request.args.get("page", 1, type=int)
    status         = request.args.get("status", "")
    pay_status     = request.args.get("pay_status", "")
    pay_method     = request.args.get("pay_method", "")
    customer_f     = request.args.get("customer", "")

    oq = Order.query.filter_by(is_deleted=False)
    if status:
        oq = oq.filter_by(order_status=status)
    if pay_status:
        oq = oq.filter_by(payment_status=pay_status)
    if pay_method:
        oq = oq.filter_by(payment_method=pay_method)
    if customer_f:
        oq = oq.filter(or_(
            Order.customer_name.ilike(f"%{customer_f}%"),
            Order.customer_phone.ilike(f"%{customer_f}%"),
        ))

    orders      = oq.order_by(Order.created_at.desc()).paginate(page=page, per_page=20)
    trash_count = Order.query.filter_by(is_deleted=True).count()
    return render_template("admin/orders.html", orders=orders,
                           status=status, pay_status=pay_status,
                           pay_method=pay_method, customer_f=customer_f,
                           trash_count=trash_count)


@app.route("/admin/orders/export-csv")
@admin_required
def export_orders_csv():
    """Download all orders as CSV — payment received/pending + order status for temple records."""
    from flask import Response

    status_filter   = request.args.get("status", "")
    payment_filter  = request.args.get("payment_status", "")

    oq = Order.query.filter_by(is_deleted=False)
    if status_filter:
        oq = oq.filter_by(order_status=status_filter)
    if payment_filter:
        oq = oq.filter_by(payment_status=payment_filter)

    all_orders = oq.order_by(Order.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Order #", "Date", "Customer Name", "Phone", "Email",
        "Address", "City", "State", "Pincode",
        "Books Ordered",
        "Subtotal (INR)", "Shipping (INR)", "Discount (INR)", "Total (INR)",
        "Payment Method", "Payment Status",
        "Order Status", "Coupon Code", "Notes"
    ])

    for order in all_orders:
        books_list = "; ".join(
            f"{item.book_title} x{item.quantity}" for item in order.items
        )
        writer.writerow([
            order.order_number,
            order.created_at.strftime("%d-%m-%Y %H:%M"),
            order.customer_name,
            order.customer_phone,
            order.customer_email or "",
            order.address,
            order.city or "",
            order.state or "",
            order.pincode or "",
            books_list,
            int(order.subtotal),
            int(order.shipping_charge),
            int(order.discount_amount),
            int(order.total_amount),
            order.payment_method.upper(),
            order.payment_status.upper(),
            order.order_status.capitalize(),
            order.coupon_code or "",
            order.notes or "",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    if status_filter:
        filename = f"orders_{status_filter}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    if payment_filter:
        filename = f"orders_payment_{payment_filter}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"

    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/admin/orders/<int:order_id>")
@admin_required
def admin_order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    return render_template("admin/order_detail.html", order=order)


@app.route("/admin/orders/update/<int:order_id>", methods=["POST"])
@admin_required
def admin_update_order(order_id):
    order               = Order.query.get_or_404(order_id)
    prev_order_status   = order.order_status
    prev_payment_status = order.payment_status

    new_order_status    = request.form.get("order_status", order.order_status)
    new_payment_status  = request.form.get("payment_status", order.payment_status)

    order.order_status   = new_order_status
    order.payment_status = new_payment_status
    order.courier_name   = request.form.get("courier_name", "").strip() or None
    order.tracking_number = request.form.get("tracking_number", "").strip() or None
    exp_del = request.form.get("expected_delivery", "").strip()
    if exp_del:
        try:
            from datetime import date
            order.expected_delivery = date.fromisoformat(exp_del)
        except ValueError:
            pass
    else:
        order.expected_delivery = None
    db.session.commit()

    # Respect the admin's "Send email notification" checkbox
    notify = request.form.get("send_email") == "1"

    if notify:
        try:
            do_confirm = False
            do_ship    = False
            do_deliver = False

            if prev_order_status != new_order_status:
                if new_order_status == "confirmed":
                    do_confirm = True
                elif new_order_status == "shipped":
                    do_ship = True
                elif new_order_status == "delivered":
                    do_deliver = True

            # Also fire confirmation when admin manually marks payment as Paid
            # (covers UPI/manual cases where order_status was already "confirmed")
            if (new_payment_status == "paid"
                    and prev_payment_status != "paid"
                    and new_order_status in ("placed", "confirmed")
                    and not do_confirm):
                do_confirm = True

            if do_confirm:
                send_order_confirmation(order)
                if order.customer_email:
                    flash(f"Confirmation email queued for {order.customer_email}", "info")
            if do_ship:
                send_order_shipped(order)
                if order.customer_email:
                    flash(f"Shipped email queued for {order.customer_email}", "info")
            if do_deliver:
                send_order_delivered(order)
                if order.customer_email:
                    flash(f"Delivered email queued for {order.customer_email}", "info")
        except Exception as e:
            app.logger.error(f"[EMAIL] Unexpected error in admin_update_order: {e}")
            flash("Order saved, but email notification failed. Check logs.", "warning")

    flash("Order updated.", "success")
    return redirect(url_for("admin_order_detail", order_id=order_id))


@app.route("/admin/orders/delete/<int:order_id>", methods=["POST"])
@admin_required
def admin_delete_order(order_id):
    order = Order.query.get_or_404(order_id)
    order_number = order.order_number
    try:
        order.is_deleted = True
        db.session.commit()
        flash(f"Order {order_number} moved to Trash. Restore it from the Trash tab if needed.", "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Could not delete order: {e}", "danger")
    return redirect(url_for("admin_orders"))


@app.route("/admin/orders/<int:order_id>/delhivery/book", methods=["POST"])
@admin_required
def admin_delhivery_book(order_id):
    """Book a Delhivery Prepaid pickup for this order."""
    from delhivery import create_shipment
    order = Order.query.get_or_404(order_id)
    if order.tracking_number and order.courier_name == "Delhivery":
        flash(f"Delhivery shipment already booked: {order.tracking_number}", "info")
        return redirect(url_for("admin_order_detail", order_id=order_id))
    waybill, err = create_shipment(order)
    if waybill:
        order.courier_name    = "Delhivery"
        order.tracking_number = waybill
        order.order_status    = "shipped"
        db.session.commit()
        send_order_shipped(order)
        flash(f"Delhivery pickup booked! Waybill: {waybill}", "success")
    else:
        flash(f"Delhivery booking failed: {err}", "danger")
    return redirect(url_for("admin_order_detail", order_id=order_id))


@app.route("/admin/orders/<int:order_id>/delhivery/track")
@admin_required
def admin_delhivery_track(order_id):
    """AJAX: return live tracking JSON for this order's Delhivery waybill."""
    from delhivery import track_shipment
    order = Order.query.get_or_404(order_id)
    if not order.tracking_number:
        return jsonify({"error": "No Delhivery waybill on this order"})
    result, err = track_shipment(order.tracking_number)
    if err:
        return jsonify({"error": err})
    return jsonify(result)


@app.route("/admin/orders/<int:order_id>/delhivery/cancel", methods=["POST"])
@admin_required
def admin_delhivery_cancel(order_id):
    """Cancel a booked Delhivery shipment and revert order status."""
    from delhivery import cancel_shipment
    order = Order.query.get_or_404(order_id)
    if not order.tracking_number:
        flash("No Delhivery waybill to cancel.", "warning")
        return redirect(url_for("admin_order_detail", order_id=order_id))
    success, msg = cancel_shipment(order.tracking_number)
    if success:
        order.courier_name    = None
        order.tracking_number = None
        order.order_status    = "confirmed"
        db.session.commit()
        flash("Delhivery shipment cancelled. Order reverted to Confirmed.", "success")
    else:
        flash(f"Cancellation failed: {msg}", "danger")
    return redirect(url_for("admin_order_detail", order_id=order_id))


@app.route("/admin/orders/trash")
@admin_required
def admin_trash_orders():
    page    = request.args.get("page", 1, type=int)
    deleted = Order.query.filter_by(is_deleted=True).order_by(Order.created_at.desc()).paginate(page=page, per_page=20)
    return render_template("admin/trash_orders.html", orders=deleted)


@app.route("/admin/orders/restore/<int:order_id>", methods=["POST"])
@admin_required
def admin_restore_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.is_deleted = False
    db.session.commit()
    flash(f"Order {order.order_number} restored successfully.", "success")
    return redirect(url_for("admin_trash_orders"))


@app.route("/admin/orders/hard-delete/<int:order_id>", methods=["POST"])
@admin_required
def admin_hard_delete_order(order_id):
    order = Order.query.get_or_404(order_id)
    order_number = order.order_number
    OrderItem.query.filter_by(order_id=order.id).delete()
    db.session.delete(order)
    db.session.commit()
    flash(f"Order {order_number} permanently deleted.", "danger")
    return redirect(url_for("admin_trash_orders"))


# ── Admin: Coupons ──

@app.route("/admin/coupons")
@admin_required
def admin_coupons():
    coupons = Coupon.query.order_by(Coupon.id.desc()).all()
    return render_template("admin/coupons.html", coupons=coupons)


@app.route("/admin/coupons/add", methods=["POST"])
@admin_required
def admin_add_coupon():
    expires_str = request.form.get("expires_at", "")
    coupon = Coupon(
        code           = request.form["code"].strip().upper(),
        description    = request.form.get("description", ""),
        discount_type  = request.form.get("discount_type", "percent"),
        discount_value = float(request.form["discount_value"]),
        min_order      = float(request.form.get("min_order", 0)),
        max_discount   = float(request.form["max_discount"]) if request.form.get("max_discount") else None,
        max_uses       = int(request.form.get("max_uses", 100)),
        active         = bool(request.form.get("active")),
        expires_at     = datetime.strptime(expires_str, "%Y-%m-%d") if expires_str else None,
    )
    db.session.add(coupon)
    db.session.commit()
    flash(f'Coupon "{coupon.code}" created!', "success")
    return redirect(url_for("admin_coupons"))


@app.route("/admin/coupons/delete/<int:coupon_id>", methods=["POST"])
@admin_required
def admin_delete_coupon(coupon_id):
    coupon = Coupon.query.get_or_404(coupon_id)
    db.session.delete(coupon)
    db.session.commit()
    flash("Coupon deleted.", "info")
    return redirect(url_for("admin_coupons"))


@app.route("/admin/backup")
@admin_required
def admin_backup():
    """Download full backup as ZIP containing orders, books, and coupons CSV files."""
    from flask import Response

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:

        # ── 1. Orders CSV ──
        orders_buf = io.StringIO()
        w = csv.writer(orders_buf)
        w.writerow([
            "Order #", "Date", "Customer Name", "Phone", "Email",
            "Address", "City", "State", "Pincode", "Books Ordered",
            "Subtotal (INR)", "Shipping (INR)", "Discount (INR)", "Total (INR)",
            "Payment Method", "Payment Status", "Order Status", "Coupon Code", "Notes"
        ])
        for order in Order.query.order_by(Order.created_at.desc()).all():
            books_list = "; ".join(f"{i.book_title} x{i.quantity}" for i in order.items)
            w.writerow([
                order.order_number,
                order.created_at.strftime("%d-%m-%Y %H:%M"),
                order.customer_name, order.customer_phone, order.customer_email or "",
                order.address, order.city or "", order.state or "", order.pincode or "",
                books_list,
                int(order.subtotal), int(order.shipping_charge),
                int(order.discount_amount), int(order.total_amount),
                order.payment_method.upper(), order.payment_status.upper(),
                order.order_status.capitalize(), order.coupon_code or "", order.notes or "",
            ])
        zf.writestr("orders.csv", orders_buf.getvalue().encode("utf-8-sig"))

        # ── 2. Books CSV ──
        books_buf = io.StringIO()
        w = csv.writer(books_buf)
        w.writerow([
            "ID", "Title", "Author", "Category", "Language",
            "Price (INR)", "Original Price (INR)", "Stock", "Pages",
            "ISBN", "Publisher", "Featured", "Active", "Is eBook"
        ])
        for book in Book.query.order_by(Book.id).all():
            w.writerow([
                book.id, book.title, book.author,
                book.category.name if book.category else "",
                book.language or "", int(book.price),
                int(book.original_price) if book.original_price else "",
                book.stock, book.pages or "",
                book.isbn or "", book.publisher or "",
                "Yes" if book.featured else "No",
                "Yes" if book.active else "No",
                "Yes" if book.is_ebook else "No",
            ])
        zf.writestr("books.csv", books_buf.getvalue().encode("utf-8-sig"))

        # ── 3. Coupons CSV ──
        coupons_buf = io.StringIO()
        w = csv.writer(coupons_buf)
        w.writerow([
            "Code", "Description", "Type", "Discount Value",
            "Min Order (INR)", "Max Discount (INR)", "Max Uses", "Times Used", "Active"
        ])
        for c in Coupon.query.order_by(Coupon.id).all():
            w.writerow([
                c.code, c.description or "", c.discount_type,
                c.discount_value, int(c.min_order) if c.min_order else 0,
                int(c.max_discount) if c.max_discount else "",
                c.max_uses or "", c.used_count, "Yes" if c.active else "No",
            ])
        zf.writestr("coupons.csv", coupons_buf.getvalue().encode("utf-8-sig"))

    zip_buffer.seek(0)
    filename = f"iskcon_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    return Response(
        zip_buffer.read(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─────────────────────────────────────────────
# UTILITY ROUTES
# ─────────────────────────────────────────────

@app.route("/api/cart-count")
def api_cart_count():
    return jsonify({"count": cart_item_count()})


@app.route("/api/search-suggestions")
def search_suggestions():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    pattern = f"%{q}%"
    books = Book.query.filter(
        Book.active == True,
        Book.deleted == False,
        db.or_(Book.title.ilike(pattern), Book.author.ilike(pattern))
    ).order_by(Book.title).limit(8).all()
    results = [{"title": b.title, "author": b.author, "id": b.id} for b in books]
    return jsonify(results)


@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500


# ─────────────────────────────────────────────
# ADMIN — CUSTOMERS
# ─────────────────────────────────────────────

@app.route("/admin/customers")
@admin_required
def admin_customers():
    q    = request.args.get("q", "").strip()
    tier = request.args.get("tier", "")

    cq = Customer.query
    if q:
        pattern = f"%{q}%"
        cq = cq.filter(
            or_(Customer.name.ilike(pattern),
                Customer.email.ilike(pattern),
                Customer.phone.ilike(pattern))
        )
    if tier:
        cq = cq.filter_by(tier=tier)

    customers = cq.order_by(Customer.created_at.desc()).all()
    return render_template("admin/customers.html",
                           customers=customers,
                           q=q, tier=tier,
                           active_page="customers")


# ─────────────────────────────────────────────
# INIT DB & RUN
# ─────────────────────────────────────────────

def init_db():
    """Create tables if they don't exist, and add any missing columns."""
    with app.app_context():
        try:
            db.create_all()
            print("[OK] Database tables created.")
        except Exception as e:
            print(f"[ERROR] db.create_all() failed: {e}")

        # Add new columns to existing tables if they don't exist (safe for PostgreSQL & SQLite)
        migrations = [
            ("orders", "courier_name",       "VARCHAR(100)"),
            ("orders", "tracking_number",    "VARCHAR(100)"),
            ("orders", "expected_delivery",  "DATE"),
            ("orders", "upi_transaction_id", "VARCHAR(100)"),
            ("orders", "is_deleted",         "BOOLEAN DEFAULT FALSE"),
            ("books",  "deleted",            "BOOLEAN DEFAULT FALSE"),
            ("books",  "is_ebook",           "BOOLEAN DEFAULT FALSE"),
            ("books",  "ebook_file",         "VARCHAR(200)"),
            ("books",              "preview_file", "VARCHAR(200)"),
            ("books",              "weight_kg",    "FLOAT DEFAULT 0.5"),
            ("stock_receipts",     "batch_ref",    "VARCHAR(20)"),
            ("orders",             "customer_id",  "INTEGER"),
        ]
        for table, column, col_type in migrations:
            # Use a fresh connection per column so a failed ALTER doesn't
            # leave the transaction in an aborted state (PostgreSQL issue)
            try:
                with db.engine.connect() as conn:
                    conn.execute(db.text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
                    conn.commit()
                    print(f"[MIGRATE] Added column {table}.{column}")
            except Exception:
                pass  # Column already exists — ignore

        # Backfill any NULL is_deleted values so filter_by(is_deleted=False) works correctly
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("UPDATE orders SET is_deleted = FALSE WHERE is_deleted IS NULL"))
                conn.commit()
        except Exception:
            pass


@app.route("/admin/export-data")
def admin_export_data():
    """
    Export all books and categories as JSON for local sync.
    Protected by a simple token (not admin session — needed for script access).
    Usage: GET /admin/export-data?token=iskcon-sync-2024
    """
    token = request.args.get("token", "")
    if token != "iskcon-sync-2024":
        abort(403)

    categories = []
    for c in Category.query.order_by(Category.sort_order).all():
        categories.append({
            "id": c.id,
            "name": c.name,
            "slug": c.slug,
            "description": c.description or "",
            "icon": c.icon or "📚",
            "sort_order": c.sort_order,
        })

    books = []
    for b in Book.query.all():
        books.append({
            "id": b.id,
            "title": b.title,
            "author": b.author,
            "description": b.description or "",
            "short_desc": b.short_desc or "",
            "price": b.price,
            "original_price": b.original_price,
            "image": b.image or "default_book.jpg",
            "category_name": b.category.name if b.category else None,
            "isbn": b.isbn or "",
            "language": b.language or "English",
            "pages": b.pages,
            "publisher": b.publisher or "The Bhaktivedanta Book Trust",
            "stock": b.stock,
            "featured": b.featured,
            "active": b.active,
            "deleted": b.deleted,
            "is_ebook": b.is_ebook,
            "ebook_file": b.ebook_file or "",
            "preview_file": b.preview_file or "",
        })

    return jsonify({"categories": categories, "books": books})


@app.route("/admin/carousel", methods=["GET", "POST"])
@admin_required
def admin_carousel():
    all_books = Book.query.filter_by(active=True).order_by(Book.title).all()
    keys = ["carousel_slide_0", "carousel_slide_1", "carousel_slide_2", "carousel_slide_3", "carousel_slide_4"]
    labels = ["Slide 0 — Welcome (brand image)", "Slide 1 — Bhagavad Gita", "Slide 2 — Srimad Bhagavatam", "Slide 3 — Krishna Book", "Slide 4 — Nectar of Devotion"]

    if request.method == "POST":
        for key in keys:
            val = request.form.get(key, "")
            Setting.set(key, val if val else None)
        flash("Carousel books updated successfully.", "success")
        return redirect(url_for("admin_carousel"))

    current = {key: Setting.get(key) for key in keys}
    return render_template("admin/carousel.html",
                           all_books=all_books,
                           keys=keys,
                           labels=labels,
                           current=current,
                           active_page="carousel")


# Auto-init DB when loaded by gunicorn
try:
    init_db()
except Exception as e:
    print(f"[ERROR] init_db failed at startup: {e}")



if __name__ == "__main__":
    init_db()
    debug = os.environ.get("FLASK_ENV", "development") == "development"
    port  = int(os.environ.get("PORT", 5000))
    host  = os.environ.get("HOST", "0.0.0.0")
    print(f"\n[START] ISKCON Book Store running at http://{host}:{port}")
    print(f"   Admin panel: http://{host}:{port}/admin/\n")
    app.run(host=host, port=port, debug=debug)
